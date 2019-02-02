#!/usr/bin/python
# -*- coding: UTF-8 -*-
# owner zhouyong0701 2018/10/10

from openpyxl import load_workbook
import getopt
import sys


def parser(source, template,target):
    print("parser source = " + source + " target = " + target)
    sourceWb = load_workbook(source)
    sourceSheet1 = sourceWb.get_sheet_by_name("表一")
    sourceSheet2 = sourceWb.get_sheet_by_name("表二")
    targetWb = load_workbook(template)
    targetSheet1 = targetWb.get_sheet_by_name("表一")
    targetSheet2 = targetWb.get_sheet_by_name("表二")
    for i in range(11,23):
        print(sourceSheet1['A' + str(i)].value)
        targetSheet1['A' + str(i-1)] = sourceSheet1['A' + str(i)].value
        targetSheet1['B' + str(i-1)] = sourceSheet1['B' + str(i)].value
        targetSheet1['C' + str(i-1)] = sourceSheet1['C' + str(i)].value
        targetSheet1['D' + str(i-1)] = sourceSheet1['D' + str(i)].value
        targetSheet1['E' + str(i-1)] = sourceSheet1['E' + str(i)].value
        targetSheet1['F' + str(i-1)] = sourceSheet1['F' + str(i)].value
        targetSheet1['G' + str(i-1)] = sourceSheet1['G' + str(i)].value
        targetSheet1['H' + str(i-1)] = sourceSheet1['H' + str(i)].value
        targetSheet1['I' + str(i-1)] = sourceSheet1['I' + str(i)].value
        targetSheet1['J' + str(i-1)] = sourceSheet1['J' + str(i)].value
        targetSheet1['K' + str(i-1)] = sourceSheet1['K' + str(i)].value
        targetSheet1['L' + str(i-1)] = sourceSheet1['L' + str(i)].value
        targetSheet1['M' + str(i-1)] = sourceSheet1['M' + str(i)].value
        targetSheet1['N' + str(i-1)] = sourceSheet1['N' + str(i)].value
        targetSheet1['O' + str(i-1)] = sourceSheet1['O' + str(i)].value
        targetSheet1['P' + str(i-1)] = sourceSheet1['P' + str(i)].value
        targetSheet1['Q' + str(i-1)] = sourceSheet1['R' + str(i)].value
        targetSheet1['R' + str(i-1)] = sourceSheet1['S' + str(i)].value
        targetSheet1['S' + str(i-1)] = sourceSheet1['T' + str(i)].value
        targetSheet1['T' + str(i-1)] = sourceSheet1['U' + str(i)].value
        targetSheet1['U' + str(i-1)] = sourceSheet1['V' + str(i)].value
        targetSheet1['V' + str(i-1)] = sourceSheet1['W' + str(i)].value
        targetSheet1['W' + str(i-1)] = sourceSheet1['X' + str(i)].value
        targetSheet1['X' + str(i-1)] = sourceSheet1['Y' + str(i)].value
        targetSheet1['Y' + str(i-1)] = sourceSheet1['Z' + str(i)].value
        targetSheet1['Z' + str(i-1)] = sourceSheet1['AA' + str(i)].value
        targetSheet1['AA' + str(i-1)] = sourceSheet1['AB' + str(i)].value
        targetSheet1['AB' + str(i-1)] = sourceSheet1['AC' + str(i)].value

    print("i = " + str(i))
    targetSheet1['B' + str(i - 1)] = '=SUM(B10:B20)'
    targetSheet1['C' + str(i - 1)] = '=SUM(C10:C20)'
    targetSheet1['D' + str(i - 1)] = '=SUM(D10:D20)'
    targetSheet1['E' + str(i - 1)] = '=SUM(E10:E20)'
    targetSheet1['F' + str(i - 1)] = '=SUM(F10:F20)'
    targetSheet1['G' + str(i - 1)] = '=SUM(G10:G20)'
    targetSheet1['H' + str(i - 1)] = '=SUM(H10:H20)'
    targetSheet1['I' + str(i - 1)] = '=SUM(I10:I20)'
    targetSheet1['J' + str(i - 1)] = '=SUM(J10:J20)'
    targetSheet1['K' + str(i - 1)] = '=SUM(K10:K20)'
    targetSheet1['L' + str(i - 1)] = '=SUM(L10:L20)'
    targetSheet1['M' + str(i - 1)] = '=SUM(M10:M20)'
    targetSheet1['N' + str(i - 1)] = '=SUM(N10:N20)'
    targetSheet1['O' + str(i - 1)] = '=SUM(O10:O20)'
    targetSheet1['P' + str(i - 1)] = '=SUM(P10:P20)'
    targetSheet1['Q' + str(i - 1)] = '=SUM(Q10:Q20)'
    targetSheet1['R' + str(i - 1)] = '=SUM(R10:R20)'
    targetSheet1['S' + str(i - 1)] = '=SUM(S10:S20)'
    targetSheet1['T' + str(i - 1)] = '=SUM(T10:T20)'
    targetSheet1['U' + str(i - 1)] = '=SUM(U10:U20)'
    targetSheet1['V' + str(i - 1)] = '=SUM(V10:V20)'
    targetSheet1['W' + str(i - 1)] = '=SUM(W10:W20)'
    targetSheet1['X' + str(i - 1)] = '=SUM(X10:X20)'
    targetSheet1['Y' + str(i - 1)] = '=SUM(Y10:Y20)'
    targetSheet1['Z' + str(i - 1)] = '=SUM(Z10:Z20)'
    targetSheet1['AA' + str(i - 1)] = '=SUM(AA10:AA20)'
    targetSheet1['AB' + str(i - 1)] = '=SUM(AB10:AB20)'

    for i in range(11,23):
        print(sourceSheet2['A' + str(i)].value)
        targetSheet2['A' + str(i-1)] = sourceSheet2['C' + str(i)].value
        targetSheet2['B' + str(i-1)] = sourceSheet2['D' + str(i)].value
        targetSheet2['C' + str(i-1)] = sourceSheet2['E' + str(i)].value
        targetSheet2['D' + str(i-1)] = sourceSheet2['F' + str(i)].value
        targetSheet2['E' + str(i-1)] = sourceSheet2['G' + str(i)].value
        targetSheet2['F' + str(i-1)] = sourceSheet2['H' + str(i)].value
        targetSheet2['G' + str(i-1)] = sourceSheet2['I' + str(i)].value
        targetSheet2['H' + str(i-1)] = sourceSheet2['J' + str(i)].value
        targetSheet2['I' + str(i-1)] = sourceSheet2['K' + str(i)].value

        targetSheet2['J' + str(i-1)] = sourceSheet2['O' + str(i)].value
        targetSheet2['K' + str(i-1)] = sourceSheet2['P' + str(i)].value
        targetSheet2['L' + str(i-1)] = sourceSheet2['Q' + str(i)].value
        targetSheet2['M' + str(i-1)] = sourceSheet2['R' + str(i)].value
        targetSheet2['N' + str(i-1)] = sourceSheet2['S' + str(i)].value
        targetSheet2['O' + str(i-1)] = sourceSheet2['T' + str(i)].value
        targetSheet2['P' + str(i-1)] = sourceSheet2['U' + str(i)].value
        targetSheet2['Q' + str(i-1)] = sourceSheet2['V' + str(i)].value
        targetSheet2['R' + str(i-1)] = sourceSheet2['W' + str(i)].value
        targetSheet2['S' + str(i-1)] = sourceSheet2['X' + str(i)].value
        targetSheet2['T' + str(i-1)] = sourceSheet2['Y' + str(i)].value
        targetSheet2['U' + str(i-1)] = sourceSheet2['Z' + str(i)].value

        targetSheet2['V' + str(i-1)] = sourceSheet2['AA' + str(i)].value
        targetSheet2['W' + str(i-1)] = sourceSheet2['AB' + str(i)].value

    targetSheet2['B' + str(i - 1)] = '=SUM(B10:B20)'
    targetSheet2['C' + str(i - 1)] = '=SUM(C10:C20)'
    targetSheet2['D' + str(i - 1)] = '=SUM(D10:D20)'
    targetSheet2['E' + str(i - 1)] = '=SUM(E10:E20)'
    targetSheet2['F' + str(i - 1)] = '=SUM(F10:F20)'
    targetSheet2['G' + str(i - 1)] = '=SUM(G10:G20)'
    targetSheet2['H' + str(i - 1)] = '=SUM(H10:H20)'
    targetSheet2['I' + str(i - 1)] = '=SUM(I10:I20)'
    targetSheet2['J' + str(i - 1)] = '=SUM(J10:J20)'
    targetSheet2['K' + str(i - 1)] = '=SUM(K10:K20)'
    targetSheet2['L' + str(i - 1)] = '=SUM(L10:L20)'
    targetSheet2['M' + str(i - 1)] = '=SUM(M10:M20)'
    targetSheet2['N' + str(i - 1)] = '=SUM(N10:N20)'
    targetSheet2['O' + str(i - 1)] = '=SUM(O10:O20)'
    targetSheet2['P' + str(i - 1)] = '=SUM(P10:P20)'
    targetSheet2['Q' + str(i - 1)] = '=SUM(Q10:Q20)'
    targetSheet2['R' + str(i - 1)] = '=SUM(R10:R20)'
    targetSheet2['S' + str(i - 1)] = '=SUM(S10:S20)'
    targetSheet2['T' + str(i - 1)] = '=SUM(T10:T20)'
    targetSheet2['U' + str(i - 1)] = '=SUM(U10:U20)'
    targetSheet2['V' + str(i - 1)] = '=SUM(V10:V20)'
    targetSheet2['W' + str(i - 1)] = '=SUM(W10:W20)'

    targetWb.save(target)
    pass


def main(argv):
    parser(source=r"D:\mnt\ex\2018年妇保新表2.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表2.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表3.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表3.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表4.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表4.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表5.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表5.xlsx")
    #parser(source=r"D:\mnt\ex\2018年妇保新表6.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表6.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表7.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表7.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表8.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表8.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表9.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表9.xlsx")
    parser(source=r"D:\mnt\ex\2018年妇保新表10.xlsx", template=r"D:\mnt\ex\2018版新表.xlsx",target=r"D:\mnt\ex\new-2018年妇保新表10.xlsx")

if __name__ == '__main__':
    main(sys.argv[1:])