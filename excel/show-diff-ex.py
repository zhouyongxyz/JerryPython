#!/usr/bin/python
# -*- coding: UTF-8 -*-
# owner zhouyong0701 2018/10/10

from openpyxl import load_workbook
import getopt
import sys


def parser(source, target):
    print("parser source = " + source + " target = " + target)
    countSource = 0;
    countTarget = 0;
    countDiff = 0;
    wbSource = load_workbook(source)
    sheetSource = wbSource.active
    wbTarget = load_workbook(target)
    sheetTarget = wbTarget.active
    for row in sheetTarget.rows:
        exist = False
        countTarget += 1
        for rowS in sheetSource.rows:
            if countTarget == 1:
                countSource += 1;
            if rowS[6].value == row[5].value:
                exist = True

        if exist == False:
            countDiff += 1
            print("编号 = " + str(row[0].value) + " row[5] = " + str(row[5].value) + " row[2] = " + str(row[2].value))

    print("Total Source Lines : " + str(countSource))
    print("Total Target Lines : " + str(countTarget))
    print("Total Diff Lines : " + str(countDiff))
    pass

def parser2(source, target):
    print("parser2 source = " + source + " target = " + target)
    countSource = 0;
    countTarget = 0;
    countDiff = 0;
    wbSource = load_workbook(source)
    sheetSource = wbSource.active
    wbTarget = load_workbook(target)
    sheetTarget = wbTarget.active
    for row in sheetSource.rows:
        exist = False
        countSource += 1
        for rowT in sheetTarget.rows:
            if countSource == 1:
                countTarget += 1
            if rowT[5].value == row[6].value:
                exist = True

        if exist == False:
            countDiff += 1
            print("编号 = " + str(row[0].value) + " row[6] = " + str(row[6].value) + " row[2] = " + str(row[2].value))

    print("Total Source Lines : " + str(countSource))
    print("Total Target Lines : " + str(countTarget))
    print("Total Diff Lines : " + str(countDiff))

    pass

def main(argv):
    try:
        opts, args = getopt.getopt(argv, "ha:i:o:", ["abi=","output=","input="])
    except getopt.GetoptError:
        print('cts-excel-report.py [-i <inputfile> -o <outputfile> -a <abi>]')
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print('cts-excel-report.py [-i <inputfile> -o <outputfile> -a <abi>]')
            sys.exit()
        elif opt in ("-i", "--input"):
            if arg != "":
                inputfile = arg
        elif opt in ("-o", "--output"):
            if arg != "":
                outputfile = arg
        elif opt in ("-a","--abi"):
            if arg != "":
                abi = arg

    parser2(source="乳腺.xlsx", target="重大  两癌筛查.xlsx")

if __name__ == '__main__':
    main(sys.argv[1:])