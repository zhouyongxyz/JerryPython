#!/usr/bin/python
# -*- coding: UTF-8 -*-
# owner zhouyong0701 2018/01/16
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
import getopt
import sys
import os



def create_excel(inputfile,outputfile):
    # moudules= tree.xpath('/Result/Module')
    # for moudule in moudules:
    #     print moudule.xpath('./@*')[0]
    wb = Workbook()
    ws = wb.create_sheet("MR1_applyPatches",0)

    ws['A1'] = "Module"
    ws['B1'] = "Patches"
    ws['C1'] = "Comments"
    ws_count = 2

    dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='A5C639', end_color='A5C639')
                            ,border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                            ,alignment=Alignment(horizontal='center',vertical='center'))
    rule = Rule(type='cellIs', dxf=dxf, formula=["10"])
    ws.conditional_formatting.add('A1:C1',rule)
    alignment = Alignment(horizontal='center',vertical='center')
    ws['A1'].alignment=alignment
    ws['B1'].alignment=alignment
    ws['C1'].alignment=alignment

    f = open(inputfile)
    line = f.readline()
    module = os.path.dirname(line)[2:]
    patches_count = 0
    alignment1 = Alignment(horizontal='left', vertical='center')
    while line:
        patch_name = line.split("/")[-1]
        cur_module = os.path.dirname(line)[2:]
        ws['B' + str(ws_count + patches_count)] = patch_name
        ws['B' + str(ws_count + patches_count)].alignment=alignment1
        patches_count += 1

        if cur_module != module:
            if patches_count > 2:
                start_cell = str(ws_count)
                end_cell = str(ws_count + patches_count - 2)
                print "merge " + start_cell + " -> " + end_cell
                ws.merge_cells('A' + start_cell + ':A' + end_cell)
            ws['A' + str(ws_count)] = module
            ws['A' + str(ws_count)].alignment=alignment1
            ws_count += patches_count - 1
            patches_count = 1
            module = cur_module

        line = f.readline()

    dxf = DifferentialStyle(font=None, fill=PatternFill(start_color='D8E4BC', end_color='D8E4BC')
                            , border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                            bottom=Side(style='thin'))
                            , alignment=Alignment(horizontal='left', vertical='center'))
    rule = Rule(type='cellIs', dxf=dxf, formula=["10"])
    ws.conditional_formatting.add('A2:C'+str(ws_count - 1), rule)
    #set the column width
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 50
    wb.save(outputfile)

def list_dir(filepath,f):
    # 遍历filepath下所有文件，包括子目录
    if os.access(filepath,os.R_OK) == False :
        return
    files = os.listdir(filepath)
    for fi in files:
        fi_d = os.path.join(filepath, fi)
        if os.path.isdir(fi_d):
            list_dir(fi_d,f)
        else:
            if file_extension(fi) == ".patch" :
                targetFile = os.path.join(filepath, fi_d)
                f.write(fi_d + "\n")
                #print fi_d
                #print fi_d.split("/")[-1]
                #print os.path.dirname(fi_d)[2:]

def file_extension(path):
  return os.path.splitext(path)[1]


def main(argv):
    abi = ''
    filepath = './'
    outputfile = './patches_report.xlsx'
    try:
        opts, args = getopt.getopt(argv, "ha:i:o:", ["abi=","output=","input="])
    except getopt.GetoptError:
        print 'cts-excel-report.py [-i <inputfile> -o <outputfile> -a <abi>]'
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print 'cts-excel-report.py [-i <inputfile> -o <outputfile> -a <abi>]'
            sys.exit()
        elif opt in ("-i", "--input"):
            if arg != "":
                inputfile = arg
        elif opt in ("-o", "--output"):
            if arg != "":
                outputfile = arg
        elif opt in ("-a","--abi"):
            if arg != "":
                abi = arg
    f = open('tmp.txt', 'w')
    list_dir(filepath,f);
    f.close()
    #create_excel("tmp.txt",outputfile)
    f = open('tmp.txt')
    line = f.readline()
    list = []
    while line:
        list.append(line)
        #print "line -> " + line
        line = f.readline()
    list.sort()
    f = open('tmp.txt', 'w')
    for item in list:
        f.write(item)
    f.close()
    create_excel("tmp.txt", outputfile)

if __name__ == '__main__':
    main(sys.argv[1:]);