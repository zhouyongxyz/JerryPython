#!/usr/bin/python
# -*- coding: UTF-8 -*-
# owner zhouyong0701 2018/01/16
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
import getopt
import sys


def create_excel(inputfile,outputfile,abi):
    tree = etree.parse(inputfile)
    # moudules= tree.xpath('/Result/Module')
    # for moudule in moudules:
    #     print moudule.xpath('./@*')[0]
    wb = Workbook()
    ws = wb.create_sheet("cts-report",0)

    ws['A1'] = "Module"
    ws['B1'] = "Passed"
    ws['C1'] = "Failed"
    ws['D1'] = "Total Tests"
    ws['E1'] = "Done"
    ws['F1'] = "Fail Case"
    ws['G1'] = "Comments"
    ws['H1'] = "Log Info"
    ws_count = 2

    dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='A5C639', end_color='A5C639')
                            ,border=Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                            ,alignment=Alignment(horizontal='center',vertical='center'))
    rule = Rule(type='cellIs', dxf=dxf, formula=["10"])
    ws.conditional_formatting.add('A1:H1',rule)
    alignment = Alignment(horizontal='center',vertical='center')
    ws['A1'].alignment=alignment
    ws['B1'].alignment=alignment
    ws['C1'].alignment=alignment
    ws['D1'].alignment=alignment
    ws['E1'].alignment=alignment
    ws['F1'].alignment=alignment
    ws['G1'].alignment=alignment
    ws['H1'].alignment=alignment

    #get all modules
    moudules = None
    if abi == '':
        moudules = tree.xpath('/Result/Module')
    else :
        #just get arm64-v8a modules
        moudules = tree.xpath('/Result/Module[@abi="'+abi+'"]')
    total = 0;
    fail = 0;
    alignment1 = Alignment(horizontal='left', vertical='center')
    #foreache moudule
    for module in moudules:
        fail_tests = module.xpath('.//Test[@result="fail"]')
        pass_tests_num = len(module.xpath('.//Test[@result="pass"]'));
        total_tests_num = len(fail_tests) + pass_tests_num
        if len(fail_tests) > 0 :
            print "parser moudule -> " + module.xpath('./@name')[0] + " fail -> " + str(len(fail_tests))
            #print etree.tostring(moudule)
            if len(fail_tests) > 1 :
                start_cell = str(ws_count)
                end_cell = str(ws_count + len(fail_tests) - 1)
                print "merge " + start_cell + " -> " + end_cell
                ws.merge_cells('A'+start_cell+':A'+end_cell)
                ws.merge_cells('B'+start_cell+':B'+end_cell)
                ws.merge_cells('C' + start_cell + ':C' + end_cell)
                ws.merge_cells('D' + start_cell + ':D' + end_cell)
                ws.merge_cells('E' + start_cell + ':E' + end_cell)
                index = ws_count
                for fail_test in fail_tests :
                    test_case_name = fail_test.xpath("../@name")[0]
                    fail_test_name = fail_test.xpath("./@name")[0]
                    fail_test_name = test_case_name + "#" + fail_test_name
                    print fail_test_name
                    ws['F'+str(index)] = fail_test_name
                    message = fail_test.xpath("./Failure/@message")[0]
                    ws['H'+str(index)] = message
                    index += 1
            else :
                test_case_name = fail_tests[0].xpath("../@name")[0]
                fail_test_name = fail_tests[0].xpath("./@name")[0]
                fail_test_name = test_case_name + "#" + fail_test_name
                print fail_test_name
                ws['F' + str(ws_count)] = fail_test_name
                message = fail_tests[0].xpath("./Failure/@message")[0]
                ws['H' + str(ws_count)] = message

            #print ws_count
            #write Module Data
            ws['A' + str(ws_count)] = module.xpath('./@abi')[0] + " " + module.xpath('./@name')[0]
            ws['A' + str(ws_count)].alignment=alignment1
            #write Passed Data
            ws['B' + str(ws_count)] = len(fail_tests)
            ws['B' + str(ws_count)].alignment = alignment
            # write Failed Data
            ws['C' + str(ws_count)] = pass_tests_num
            ws['C' + str(ws_count)].alignment = alignment
            # write Total Tests
            ws['D' + str(ws_count)] = total_tests_num
            ws['D' + str(ws_count)].alignment = alignment
            # write Done Data
            ws['E' + str(ws_count)] = module.xpath('./@done')[0]
            ws['E' + str(ws_count)].alignment = alignment
            ws_count = ws_count + len(fail_tests)
            total += 1
            fail += len(fail_tests)

    dxf = DifferentialStyle(font=None, fill=PatternFill(start_color='D8E4BC', end_color='D8E4BC')
                            , border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                            bottom=Side(style='thin'))
                            , alignment=Alignment(horizontal='left', vertical='center'))
    rule = Rule(type='cellIs', dxf=dxf, formula=["10"])
    ws.conditional_formatting.add('A2:H'+str(ws_count - 1), rule)
    #set the column width
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['F'].width = 80
    ws.column_dimensions['G'].width = 80
    ws.column_dimensions['H'].width = 80
    print "total moudules: " + str(total) + " fail : " + str(fail)
    wb.save(outputfile)

def main(argv):
    abi = ''
    inputfile = './test_result.xml'
    outputfile = './cts-excel-report.xlsx'
    try:
        opts, args = getopt.getopt(argv, "ha:i:o:", ["abi=","output=","input="])
    except getopt.GetoptError:
        print 'cts-excel-report.py [-i <inputfile> -o <outputfile> -a <abi>]'
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print 'cts-excel-report.py [-i <inputfile> -o <outputfile> -a <abi>]'
            sys.exit()
        elif opt in ("-i", "--input"):
            if arg != "":
                inputfile = arg
        elif opt in ("-o", "--output"):
            if arg != "":
                outputfile = arg
        elif opt in ("-a","--abi"):
            if arg != "":
                abi = arg

    create_excel(inputfile,outputfile,abi);


if __name__ == '__main__':
    main(sys.argv[1:]);